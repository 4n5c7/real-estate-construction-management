"""CSV / Excel インポートモジュール (基本設計書 13章)

- 工事 ID 空欄なら自動採番
- 物件/業者/担当者はマスタ突合し、無い場合は「新規候補」として返す
- 日付は YYYY/MM/DD・YYYY-MM-DD どちらも許容
- 金額のカンマ・円記号・全角数字を許容
- 全件トランザクション、1 件でも DB エラーで全件ロールバック
"""
from __future__ import annotations

import csv
import io
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from .database import STATUS_LABELS, db_cursor
from .services import _to_int, create_construction


# ヘッダー名 → 内部キーのマッピング
HEADER_MAP = {
    "工事ID": "construction_id",
    "物件名": "property_name",
    "部屋番号": "room_no",
    "受付日": "received_at",
    "工事区分": "category_name",
    "優先度": "priority",
    "状態": "status",
    "担当者": "staff_name",
    "業者": "vendor_name",
    "工事内容": "description",
    "着工予定日": "scheduled_start_at",
    "完了予定日": "scheduled_end_at",
    "完了日": "completed_at",
    "見積金額": "estimate_amount",
    "請求金額": "invoice_amount",
    "備考": "note",
}

STATUS_NAME_TO_CODE = {v: k for k, v in STATUS_LABELS.items()}
PRIORITY_NAME_TO_CODE = {"通常": 1, "急ぎ": 2, "緊急": 3}


def _norm_date(s: Any) -> Optional[str]:
    if s is None or s == "":
        return None
    if isinstance(s, datetime):
        return s.date().isoformat()
    text = str(s).strip()
    if not text:
        return None
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None  # 解釈不能


def _read_rows(file_path: Path) -> List[Dict[str, Any]]:
    """CSV / Excel のどちらでも辞書のリストとして読み込む。"""
    ext = file_path.suffix.lower()
    rows: List[Dict[str, Any]] = []
    if ext == ".csv":
        # まず UTF-8 BOM、ダメなら Shift_JIS
        text = None
        for enc in ("utf-8-sig", "utf-8", "cp932", "shift_jis"):
            try:
                with open(file_path, "r", encoding=enc, newline="") as f:
                    text = f.read()
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise ValueError("CSV の文字コードを判別できません")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            rows.append({k.strip(): (v if v is None else v.strip())
                         for k, v in row.items() if k})
    elif ext in (".xlsx", ".xlsm"):
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = {h: v for h, v in zip(headers, r) if h}
            rows.append(row)
    else:
        raise ValueError(f"未対応の拡張子: {ext}")
    return rows


def analyze_import(file_path: Path) -> Dict[str, Any]:
    """取り込み前の検証を行い、結果を返す (確定はしない)。"""
    raw_rows = _read_rows(file_path)

    # 既存マスタを名前→ID 化
    with db_cursor() as cur:
        cur.execute("SELECT property_id, name FROM properties WHERE is_active=1")
        prop_map = {r["name"]: r["property_id"] for r in cur.fetchall()}
        cur.execute("SELECT vendor_id, name FROM vendors WHERE is_active=1")
        vend_map = {r["name"]: r["vendor_id"] for r in cur.fetchall()}
        cur.execute("SELECT staff_id, name FROM staffs WHERE is_active=1")
        staff_map = {r["name"]: r["staff_id"] for r in cur.fetchall()}
        cur.execute("SELECT category_id, name FROM categories WHERE is_active=1")
        cat_map = {r["name"]: r["category_id"] for r in cur.fetchall()}

    new_properties: set = set()
    new_vendors: set = set()
    new_staffs: set = set()
    new_categories: set = set()
    errors: List[Dict[str, Any]] = []
    valid_rows: List[Dict[str, Any]] = []

    for i, row in enumerate(raw_rows, start=2):  # 行番号は2行目から
        # ヘッダーを内部キーへ
        rec: Dict[str, Any] = {}
        for jp, key in HEADER_MAP.items():
            if jp in row:
                rec[key] = row[jp]

        # 必須: 物件名
        if not rec.get("property_name"):
            errors.append({"row": i, "error": "物件名が必須です", "data": row})
            continue

        # 受付日
        rec_date = _norm_date(rec.get("received_at"))
        if rec.get("received_at") and rec_date is None:
            errors.append({"row": i, "error": "受付日の形式が不正", "data": row})
            continue
        rec["received_at"] = rec_date

        for f in ("scheduled_start_at", "scheduled_end_at", "completed_at"):
            v = _norm_date(rec.get(f))
            rec[f] = v

        # 金額
        for f in ("estimate_amount", "invoice_amount"):
            rec[f] = _to_int(rec.get(f))

        # 優先度
        p = rec.get("priority")
        if isinstance(p, str) and p in PRIORITY_NAME_TO_CODE:
            rec["priority"] = PRIORITY_NAME_TO_CODE[p]
        else:
            try:
                rec["priority"] = int(p) if p else 1
            except (ValueError, TypeError):
                rec["priority"] = 1

        # 状態
        st = rec.get("status")
        if st in STATUS_NAME_TO_CODE:
            rec["status"] = STATUS_NAME_TO_CODE[st]
        elif st in STATUS_LABELS:
            rec["status"] = st
        else:
            rec["status"] = "received"

        # マスタ突合
        rec["property_id"] = prop_map.get(rec.get("property_name"))
        if rec["property_id"] is None:
            new_properties.add(rec["property_name"])

        if rec.get("vendor_name"):
            rec["vendor_id"] = vend_map.get(rec["vendor_name"])
            if rec["vendor_id"] is None:
                new_vendors.add(rec["vendor_name"])
        if rec.get("staff_name"):
            rec["staff_id"] = staff_map.get(rec["staff_name"])
            if rec["staff_id"] is None:
                new_staffs.add(rec["staff_name"])
        if rec.get("category_name"):
            rec["category_id"] = cat_map.get(rec["category_name"])
            if rec["category_id"] is None:
                new_categories.add(rec["category_name"])

        rec["_row"] = i
        valid_rows.append(rec)

    return {
        "rows": valid_rows,
        "errors": errors,
        "new_properties": sorted(new_properties),
        "new_vendors": sorted(new_vendors),
        "new_staffs": sorted(new_staffs),
        "new_categories": sorted(new_categories),
        "total": len(raw_rows),
    }


def commit_import(analyzed: Dict[str, Any], auto_create_masters: bool = True
                  ) -> Dict[str, Any]:
    """検証済みデータを確定登録する。"""
    created_props = 0
    created_vends = 0
    created_staffs = 0
    created_cats = 0
    created_cons = 0

    # 新規マスタの作成
    if auto_create_masters:
        with db_cursor() as cur:
            for name in analyzed["new_properties"]:
                cur.execute("INSERT INTO properties(name) VALUES (?)", (name,))
                created_props += 1
            for name in analyzed["new_vendors"]:
                cur.execute("INSERT INTO vendors(name) VALUES (?)", (name,))
                created_vends += 1
            for name in analyzed["new_staffs"]:
                cur.execute("INSERT INTO staffs(name) VALUES (?)", (name,))
                created_staffs += 1
            for name in analyzed["new_categories"]:
                cur.execute("INSERT INTO categories(name) VALUES (?)", (name,))
                created_cats += 1

    # マップ再取得
    with db_cursor() as cur:
        cur.execute("SELECT property_id, name FROM properties")
        prop_map = {r["name"]: r["property_id"] for r in cur.fetchall()}
        cur.execute("SELECT vendor_id, name FROM vendors")
        vend_map = {r["name"]: r["vendor_id"] for r in cur.fetchall()}
        cur.execute("SELECT staff_id, name FROM staffs")
        staff_map = {r["name"]: r["staff_id"] for r in cur.fetchall()}
        cur.execute("SELECT category_id, name FROM categories")
        cat_map = {r["name"]: r["category_id"] for r in cur.fetchall()}

    # 案件作成
    errors: List[Dict[str, Any]] = []
    for rec in analyzed["rows"]:
        rec["property_id"] = prop_map.get(rec.get("property_name"))
        if rec.get("vendor_name"):
            rec["vendor_id"] = vend_map.get(rec["vendor_name"])
        if rec.get("staff_name"):
            rec["staff_id"] = staff_map.get(rec["staff_name"])
        if rec.get("category_name"):
            rec["category_id"] = cat_map.get(rec["category_name"])
        try:
            create_construction(rec)
            created_cons += 1
        except Exception as e:  # noqa: BLE001
            errors.append({"row": rec.get("_row"), "error": str(e)})

    return {
        "created_constructions": created_cons,
        "created_properties": created_props,
        "created_vendors": created_vends,
        "created_staffs": created_staffs,
        "created_categories": created_cats,
        "errors": errors,
    }
