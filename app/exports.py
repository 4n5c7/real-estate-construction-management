"""Excel / CSV 出力モジュール (基本設計書 12 章)

無償ライブラリ openpyxl を使用。
出力テンプレートは用途別に複数用意する。
"""
from __future__ import annotations

import csv
import io
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import get_paths, load_settings
from .database import PRIORITY_LABELS, STATUS_LABELS


HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _autosize(ws) -> None:
    """列幅の自動調整 (簡易)。"""
    for col in ws.columns:
        max_len = 8
        letter = get_column_letter(col[0].column)
        for cell in col:
            v = cell.value
            if v is None:
                continue
            l = len(str(v))
            if l > max_len:
                max_len = l
        ws.column_dimensions[letter].width = min(max_len + 2, 40)


def _write_header(ws, headers: List[str]) -> None:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")


def _yen(v) -> str:
    if v is None or v == "":
        return ""
    try:
        return f"{int(v):,}"
    except (ValueError, TypeError):
        return str(v)


# ---------------------------------------------------------------------------
# 工事一覧表 (12.2.1)
# ---------------------------------------------------------------------------
def export_construction_list(rows: Iterable[Dict[str, Any]], out_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "工事一覧"
    headers = [
        "工事ID", "物件名", "部屋番号", "工事区分", "状態", "優先度",
        "担当者", "業者", "受付日", "完了予定日", "完了日",
        "見積金額", "請求金額", "見積書", "請求書", "備考",
    ]
    _write_header(ws, headers)
    for r in rows:
        ws.append([
            r.get("construction_id"),
            r.get("property_name"),
            r.get("room_no"),
            r.get("category_name"),
            STATUS_LABELS.get(r.get("status"), r.get("status")),
            PRIORITY_LABELS.get(r.get("priority"), r.get("priority")),
            r.get("staff_name"),
            r.get("vendor_name"),
            r.get("received_at"),
            r.get("scheduled_end_at"),
            r.get("completed_at"),
            _yen(r.get("estimate_amount")),
            _yen(r.get("invoice_amount")),
            "あり" if r.get("has_estimate") else "なし",
            "あり" if r.get("has_invoice") else "なし",
            r.get("note"),
        ])
    _autosize(ws)
    ws.freeze_panes = "A2"
    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# オーナー報告用 (12.2.2)
# ---------------------------------------------------------------------------
def export_owner_report(rows: Iterable[Dict[str, Any]], out_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "オーナー報告"
    headers = [
        "物件名", "部屋番号", "工事内容", "状態", "見積金額",
        "請求金額", "完了予定日", "完了日", "備考",
    ]
    _write_header(ws, headers)
    for r in rows:
        ws.append([
            r.get("property_name"),
            r.get("room_no"),
            r.get("description"),
            STATUS_LABELS.get(r.get("status"), r.get("status")),
            _yen(r.get("estimate_amount")),
            _yen(r.get("invoice_amount")),
            r.get("scheduled_end_at"),
            r.get("completed_at"),
            r.get("note"),
        ])
    _autosize(ws)
    ws.freeze_panes = "A2"
    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# 経理確認用 (12.2.3)
# ---------------------------------------------------------------------------
def export_accounting(rows: Iterable[Dict[str, Any]], out_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "経理確認"
    headers = [
        "工事ID", "物件名", "業者", "請求金額", "請求書添付",
        "請求書受領日", "支払予定日", "支払済", "備考",
    ]
    _write_header(ws, headers)
    for r in rows:
        ws.append([
            r.get("construction_id"),
            r.get("property_name"),
            r.get("vendor_name"),
            _yen(r.get("invoice_amount")),
            "あり" if r.get("has_invoice") else "なし",
            r.get("invoice_received_at"),
            r.get("payment_due_at"),
            "済" if r.get("paid") else "未",
            r.get("note"),
        ])
    _autosize(ws)
    ws.freeze_panes = "A2"
    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# 業者別集計 (12.2.4)
# ---------------------------------------------------------------------------
def export_vendor_summary(rows: Iterable[Dict[str, Any]], out_path: Path) -> Path:
    # 業者別に集計
    agg: Dict[str, Dict[str, Any]] = {}
    for r in rows:
        name = r.get("vendor_name") or "(未指定)"
        a = agg.setdefault(name, {"count": 0, "estimate": 0, "invoice": 0, "open": 0})
        a["count"] += 1
        a["estimate"] += r.get("estimate_amount") or 0
        a["invoice"] += r.get("invoice_amount") or 0
        if r.get("status") not in ("completed", "invoiced", "cancelled"):
            a["open"] += 1

    wb = Workbook()
    ws = wb.active
    ws.title = "業者別集計"
    _write_header(ws, ["業者名", "件数", "見積合計", "請求合計", "未完了件数"])
    for name, a in sorted(agg.items()):
        ws.append([name, a["count"], _yen(a["estimate"]),
                   _yen(a["invoice"]), a["open"]])
    _autosize(ws)
    ws.freeze_panes = "A2"
    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# CSV 出力
# ---------------------------------------------------------------------------
def export_csv(rows: Iterable[Dict[str, Any]], out_path: Path) -> Path:
    settings = load_settings()
    encoding = settings.get("default_encoding", "utf-8-sig")
    headers = [
        "工事ID", "物件名", "部屋番号", "工事区分", "状態", "優先度",
        "担当者", "業者", "受付日", "完了予定日", "完了日",
        "見積金額", "請求金額", "見積書", "請求書", "備考",
    ]
    with open(out_path, "w", encoding=encoding, newline="") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for r in rows:
            w.writerow([
                r.get("construction_id"),
                r.get("property_name"),
                r.get("room_no"),
                r.get("category_name"),
                STATUS_LABELS.get(r.get("status"), r.get("status")),
                PRIORITY_LABELS.get(r.get("priority"), r.get("priority")),
                r.get("staff_name"),
                r.get("vendor_name"),
                r.get("received_at"),
                r.get("scheduled_end_at"),
                r.get("completed_at"),
                r.get("estimate_amount") or "",
                r.get("invoice_amount") or "",
                "あり" if r.get("has_estimate") else "なし",
                "あり" if r.get("has_invoice") else "なし",
                r.get("note") or "",
            ])
    return out_path


# ---------------------------------------------------------------------------
# 出力エントリポイント
# ---------------------------------------------------------------------------
TEMPLATES = {
    "list": ("工事一覧表", export_construction_list, "xlsx"),
    "owner": ("オーナー報告", export_owner_report, "xlsx"),
    "accounting": ("経理確認", export_accounting, "xlsx"),
    "vendor": ("業者別集計", export_vendor_summary, "xlsx"),
    "csv": ("CSV (工事一覧)", export_csv, "csv"),
}


def export(template_key: str, rows: List[Dict[str, Any]]) -> Path:
    if template_key not in TEMPLATES:
        raise ValueError(f"不明なテンプレート: {template_key}")
    name, func, ext = TEMPLATES[template_key]
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    paths = get_paths()
    out_path = paths["exports"] / f"{name}_{ts}.{ext}"
    return func(rows, out_path)
